"""
openpyxl-based parser for visual Tally GST Invoice Excel files.
These files are created by PDF-to-Excel converters and retain the
visual layout of the original PDF instead of a tabular structure.

The parser works by:
1. Reading all cells with their ROW and COLUMN positions.
2. Scanning the grid to find known labels (Invoice No, Dated, Buyer, etc.)
3. Looking at cells to the RIGHT or BELOW the label cell to get the value.
4. Finding the product table header row and scanning rows below it for line items.
5. Chunking the whole sheet into invoices by finding new invoice start markers.
"""

import io
import re
import time
import json
import pandas as pd

try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from google import genai
    from google.genai import types as genai_types
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────────────────────
# PDF READING
# ─────────────────────────────────────────────────────────────

def _read_pdf_pages(file_bytes):
    """Extract text page-by-page. Returns ([(page_num, text), ...], error)."""
    if not PDF_AVAILABLE:
        return [], "pdfplumber not installed."
    try:
        pages = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                return [], "Empty PDF."
            for i, page in enumerate(pdf.pages):
                t = page.extract_text(x_tolerance=3, y_tolerance=3)
                if t and t.strip():
                    pages.append((i + 1, t.strip()))
        if not pages:
            return [], "Scanned PDF — no extractable text. Use text-based Tally export."
        return pages, None
    except Exception as e:
        return [], f"PDF read error: {str(e)}"


# ─────────────────────────────────────────────────────────────
# GEMINI AI
# ─────────────────────────────────────────────────────────────

GEMINI_PROMPT = """You are a data extraction assistant for Indian GST invoices.

Given the raw text extracted from a PDF containing ONE OR MORE Tally GST Invoices, extract the details for EACH invoice.
Return ONLY a valid JSON array of objects, with no markdown formatting and no explanation.

Required JSON structure:
[
  {
    "date": "DD-Mon-YY or DD/MM/YYYY string, or null",
    "invoice_no": "invoice number string, or null",
    "customer": "exact buyer company name from Buyer/Bill to section ONLY — NOT the seller, NOT vehicle number",
    "products": [
      {
        "description": "product or item description",
        "quantity": numeric_or_null,
        "unit": "NOS/PCS/KG/MTR/SET or empty string",
        "amount": numeric_line_item_amount_in_rupees_or_null
      }
    ]
  }
]

Critical rules:
- Extract EVERY invoice you find in the text as a separate object in the array.
- customer = the BUYER party only (comes after Buyer/Bill to label).
- EXCLUDE: SGST, CGST, IGST, Output Tax, Total rows from products.
- EXCLUDE: vehicle numbers (e.g. MH12CT1998), dispatch info, delivery terms from customer name.
- EXCLUDE: sub-description italic lines below main product.
- quantity = physical unit count (e.g. 30, 1000), NOT rupee amounts.
- amount = line item value in rupees (e.g. 9300.00, 12500.00).
- If no products found for an invoice, set products to [].
- Return ONLY the JSON array.

Text to parse:
"""

# Models available on this advanced API key
GEMINI_MODELS = [
    'gemini-2.5-flash',
    'gemini-2.0-flash-lite',
    'gemini-2.0-flash'
]


def _gemini_parse_all(text, api_key):
    """Parse entire text via Gemini in ONE call. Returns (list_of_dicts, error)."""
    try:
        client = genai.Client(api_key=api_key)
    except Exception as e:
        return None, f"Failed to init Gemini client: {e}"

    last_err = "No models succeeded"
    for model_name in GEMINI_MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=GEMINI_PROMPT + text,
                config=genai_types.GenerateContentConfig(
                    temperature=0,
                    max_output_tokens=8192,
                    response_mime_type="application/json",
                )
            )
            raw = response.text.strip()

            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError:
                raw_clean = re.sub(r'```(?:json)?\s*', '', raw)
                raw_clean = re.sub(r'\s*```', '', raw_clean)
                raw_clean = re.sub(r',\s*([\]}])', r'\1', raw_clean)
                try:
                    parsed = json.loads(raw_clean)
                except json.JSONDecodeError as nested_e:
                    err_pos = getattr(nested_e, 'pos', 0)
                    snippet = raw_clean[max(0, err_pos-40):min(len(raw_clean), err_pos+40)]
                    return None, f"JSON format error from {model_name}: {nested_e}. Snippet: '...{snippet}...' "

            if isinstance(parsed, dict):
                parsed = [parsed]
            return parsed, None

        except Exception as e:
            err_str = str(e)
            if '429' in err_str or 'quota' in err_str.lower():
                last_err = f"{model_name}: Quota exceeded"
                continue
            if '404' in err_str or 'not found' in err_str.lower():
                last_err = f"{model_name}: Model not found"
                continue
            last_err = f"API error ({model_name}): {err_str}"

    return None, last_err


# ─────────────────────────────────────────────────────────────
# MAIN PARSER
# ─────────────────────────────────────────────────────────────

SKIP_PRODUCT_KEYWORDS = [
    'sgst', 'cgst', 'igst', 'output cgst', 'output sgst',
    'total', 'grand total', 'amount in words', 'taxable value',
    'central tax', 'state tax', 'output tax'
]

def parse_gst_invoice_pdf(file_bytes, api_key=None):
    """
    Parse a Tally GST Invoice PDF (single or merged multi-invoice).
    Sends ALL text to Gemini in ONE request to avoid rate limits.
    Returns (list_of_row_dicts, error_string).
    """
    pages, err = _read_pdf_pages(file_bytes)
    if err:
        return [], err
    if not pages:
        return [], "Scanned PDF — no extractable text."

    full_text = '\n'.join(t for _, t in pages)

    # ── Gemini path: process in chunks to prevent JSON breakage ──
    if api_key and GEMINI_AVAILABLE:
        chunk_size = 5
        all_rows = []
        page_errors = []

        for i in range(0, len(pages), chunk_size):
            chunk_pages = pages[i:i + chunk_size]
            chunk_text = '\n'.join(t for _, t in chunk_pages)

            invoices, perr = _gemini_parse_all(chunk_text, api_key)
            if perr:
                start_p = chunk_pages[0][0]
                end_p = chunk_pages[-1][0]
                page_errors.append(f"Pages {start_p}-{end_p}: {perr}")
                continue

            if not invoices:
                continue

            for inv in invoices:
                customer   = str(inv.get('customer') or 'UNKNOWN').strip()
                date_val   = inv.get('date')
                invoice_no = str(inv.get('invoice_no') or 'UNKNOWN').strip()
                products   = inv.get('products') or []

                for p in products:
                    desc = str(p.get('description') or '').strip()
                    if not desc or len(desc) < 3:
                        continue
                    if any(kw in desc.lower() for kw in SKIP_PRODUCT_KEYWORDS):
                        continue
                    all_rows.append({
                        'date':       date_val,
                        'invoice_no': invoice_no,
                        'customer':   customer,
                        'product':    desc,
                        'quantity':   p.get('quantity'),
                        'unit':       str(p.get('unit') or ''),
                        'amount':     p.get('amount'),
                    })

            # Delay between chunks to avoid rate limits
            if i + chunk_size < len(pages):
                time.sleep(2)

        if not all_rows and page_errors:
            return [], f"AI parse failed: {'; '.join(page_errors)}"
        elif not all_rows:
            return [], "AI parsed invoices but found 0 valid products."

        return all_rows, None

    # ── Regex fallback (no API key) ─────────────────────────────
    return _regex_parse_text(full_text)


# ─────────────────────────────────────────────────────────────
# VISUAL EXCEL INVOICE READING (openpyxl cell-position aware)
# ─────────────────────────────────────────────────────────────

def _cell_str(cell_value):
    """Clean cell value to string."""
    if cell_value is None:
        return ''
    s = str(cell_value).strip()
    # Remove float-like .0 suffix
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    return s


def _is_likely_date(s):
    """Return True if string looks like a date."""
    return bool(re.match(
        r'^\d{1,2}[-/]\w{2,3}[-/]\d{2,4}$|^\d{1,2}[-/]\d{1,2}[-/]\d{4}$',
        s.strip(), re.IGNORECASE
    ))


def _is_likely_number(s):
    """Return True if string is a pure number."""
    return bool(re.match(r'^[\d,]+\.?\d*$', s.strip()))


SKIP_CUSTOMER_WORDS = {
    'gstin', 'gstin/uin', 'state', 'state name', 'code', 'contact',
    'e-mail', 'email', 'udyam', 'iso', 'certified', 'dated', 'voucher',
    'mode', 'payment', 'buyer', 'bill to', 'dispatch', 'destination',
    'vehicle', 'lading', 'delivery', 'term', 'place', 'supply', 'road',
    'declaration', 'ref', 'order', 'signature', 'authorised', 'bank',
    'account', 'ifsc', 'branch', 'dc no', 'contact person'
}


def parse_excel_invoice(file_bytes):
    """
    Parse a visual Tally GST Invoice Excel file (converted from PDF).
    Uses openpyxl to read exact cell positions, finding labeled fields 
    and table rows directly from the grid.
    Returns (list_of_row_dicts, error_string).
    """
    if not OPENPYXL_AVAILABLE:
        return [], "openpyxl not installed. Run: pip install openpyxl"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return [], f"Could not open Excel: {e}"

    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Build a 2D dict: {row: {col: value_str}}
        grid = {}
        for row in ws.iter_rows():
            for cell in row:
                v = _cell_str(cell.value)
                if v:
                    grid.setdefault(cell.row, {})[cell.column] = v

        if not grid:
            continue

        max_row = max(grid.keys())

        # Find all invoice start rows: rows containing "GST INVOICE" or "TAX INVOICE"
        invoice_start_rows = []
        for r, cols in grid.items():
            row_text = ' '.join(cols.values()).lower()
            if 'gst invoice' in row_text or 'tax invoice' in row_text:
                invoice_start_rows.append(r)

        if not invoice_start_rows:
            # Fallback: treat the whole sheet as one invoice starting at row 1
            invoice_start_rows = [min(grid.keys())]

        # Add sentinel end
        invoice_end_rows = invoice_start_rows[1:] + [max_row + 1]

        for inv_start, inv_end in zip(invoice_start_rows, invoice_end_rows):
            rows_parsed = _parse_one_invoice_from_grid(grid, inv_start, inv_end)
            all_rows.extend(rows_parsed)

    if not all_rows:
        return [], "No product rows extracted. Check the file format or use PDF upload with AI."

    return all_rows, None


def _parse_one_invoice_from_grid(grid, start_row, end_row):
    """Parse one invoice block from the cell grid."""

    # ── Step 1: Extract header fields ────────────────────────
    date_val   = None
    invoice_no = None
    customer   = None

    DATE_LABELS    = {'dated', 'date', 'invoice date', 'bill date'}
    INV_NO_LABELS  = {'invoice no', 'invoice no.', 'voucher no', 'bill no', 'inv no', 'inv. no.'}
    BUYER_LABELS   = {'buyer', 'buyer (bill to)', 'bill to', 'consignee', 'buyer/bill to', 'ship to'}

    for r in range(start_row, end_row):
        row_data = grid.get(r, {})
        for col, val in row_data.items():
            vl = val.lower().strip().rstrip('.:')

            # Date
            if vl in DATE_LABELS or 'dated' == vl:
                # Value is to the right
                for dc in range(col + 1, col + 6):
                    candidate = row_data.get(dc, '')
                    if candidate and _is_likely_date(candidate):
                        date_val = candidate
                        break
                # or on same row far right
                if not date_val:
                    for dc in sorted(row_data.keys()):
                        if dc > col and _is_likely_date(row_data[dc]):
                            date_val = row_data[dc]
                            break

            # Invoice number
            if vl in INV_NO_LABELS or 'invoice no' in vl:
                for dc in range(col + 1, col + 8):
                    candidate = row_data.get(dc, '')
                    if candidate and not _is_likely_date(candidate) and len(candidate) > 2:
                        invoice_no = candidate
                        break

            # Buyer/Customer name — look at cells from col+1 down for a few rows
            if vl in BUYER_LABELS or 'bill to' in vl or vl.startswith('buyer'):
                # Often the company name is a few rows below and starts in same or nearby col
                for dr in range(r, min(r + 12, end_row)):
                    next_row_data = grid.get(dr, {})
                    for dc in sorted(next_row_data.keys()):
                        candidate = next_row_data.get(dc, '').strip()
                        cl = candidate.lower()
                        # Skip headers, short strings, GSTIN patterns, dates & numbers
                        if len(candidate) < 5:
                            continue
                        if _is_likely_date(candidate) or _is_likely_number(candidate):
                            continue
                        if any(skip in cl for skip in SKIP_CUSTOMER_WORDS):
                            continue
                        if re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}', candidate):
                            continue  # GSTIN
                        if re.match(r'^[A-Z]{2}\d{2}[A-Z]{2}\d{4}$', candidate):
                            continue  # Vehicle number
                        # Accept this as customer name
                        customer = candidate
                        break
                    if customer:
                        break

    # ── Step 2: Find the product table ───────────────────────
    PROD_HDR_KW = ['description', 'goods', 'hsn', 'qty', 'quantity', 'amount', 'rate']
    prod_hdr_row = None
    qty_col   = None
    desc_col  = None
    amt_col   = None
    unit_col  = None
    rate_col  = None
    hsn_col   = None

    for r in range(start_row, end_row):
        row_data = grid.get(r, {})
        row_text = ' '.join(row_data.values()).lower()
        hits = sum(1 for kw in PROD_HDR_KW if kw in row_text)
        if hits >= 3:
            prod_hdr_row = r
            # Map columns
            for col, val in row_data.items():
                vl = val.lower().strip()
                if 'description' in vl or 'goods' in vl:
                    desc_col = col
                elif 'hsn' in vl or 'sac' in vl:
                    hsn_col = col
                elif 'qty' in vl or 'quantity' in vl:
                    qty_col = col
                elif 'rate' in vl:
                    rate_col = col
                elif 'amount' in vl or 'value' in vl:
                    amt_col = col
                elif 'unit' in vl or 'per' in vl or vl in ('nos', 'pcs', 'kg', 'mtr'):
                    unit_col = col
            break

    items = []
    if prod_hdr_row is not None:
        for r in range(prod_hdr_row + 1, end_row):
            row_data = grid.get(r, {})
            if not row_data:
                continue

            # Get description
            desc = ''
            if desc_col and desc_col in row_data:
                desc = row_data[desc_col].strip()
            else:
                # Take the leftmost non-numeric cell
                for col in sorted(row_data.keys()):
                    v = row_data[col].strip()
                    if v and not _is_likely_number(v) and not _is_likely_date(v) and len(v) > 3:
                        desc = v
                        break

            if not desc or len(desc) < 3:
                continue

            # Skip header kw rows, tax rows, total rows
            desc_lower = desc.lower()
            if any(kw in desc_lower for kw in SKIP_PRODUCT_KEYWORDS):
                continue
            if desc_lower in {'sl no', 'sl.no', 'sno', 's.no', 'sr no', '#'}:
                continue
            # Skip pure numeric (row numbers)
            if re.match(r'^\d+$', desc):
                continue

            # Get quantity
            qty = None
            if qty_col and qty_col in row_data:
                m = re.search(r'(\d+(?:\.\d+)?)', row_data[qty_col])
                if m:
                    qty = float(m.group(1))

            # Get unit
            unit = ''
            if unit_col and unit_col in row_data:
                unit = row_data[unit_col].strip().upper()
            else:
                # Scan row for unit keyword
                for col, v in row_data.items():
                    if re.match(r'^(NOS|PCS|KG|MTR|SET|BOX|LTR|MTS|UNIT)\.?$', v.strip(), re.IGNORECASE):
                        unit = v.strip().upper().rstrip('.')
                        break

            # Get amount
            amount = None
            if amt_col and amt_col in row_data:
                m = re.search(r'([\d,]+\.\d{2})', row_data[amt_col].replace(',', ''))
                if m:
                    amount = float(m.group(1))
            else:
                # Take the rightmost decimal number in the row
                nums = []
                for col in sorted(row_data.keys(), reverse=True):
                    v = row_data[col]
                    m = re.match(r'^([\d,]+\.\d{2})$', v.strip())
                    if m:
                        nums.append(float(m.group(1).replace(',', '')))
                if nums:
                    amount = nums[0]

            if desc and amount:
                items.append({
                    'date':       date_val,
                    'invoice_no': invoice_no or 'UNKNOWN',
                    'customer':   customer or 'UNKNOWN',
                    'product':    desc,
                    'quantity':   qty,
                    'unit':       unit,
                    'amount':     amount,
                })

    return items


# ─────────────────────────────────────────────────────────────
# TEXT REGEX FALLBACK (for PDFs with no API key)
# ─────────────────────────────────────────────────────────────

def _regex_parse_text(text):
    """Best-effort regex parser for PDF text when no Gemini key is set."""
    lines = text.splitlines()

    HEADER_KW = ['description', 'goods', 'hsn', 'sac', 'qty', 'quantity', 'amount', 'rate']
    hdr_indices = [i for i, l in enumerate(lines)
                   if sum(1 for kw in HEADER_KW if kw in l.lower()) >= 3]

    if not hdr_indices:
        return [], "No product rows found. Set GEMINI_API_KEY for AI parsing."

    # Split at invoice starts (look for "GST INVOICE" or "Buyer" keywords between headers)
    split_indices = []
    for i, l in enumerate(lines):
        ll = l.lower()
        if 'gst invoice' in ll or 'tax invoice' in ll:
            split_indices.append(i)

    if not split_indices:
        split_indices = [0]
    # Deduplicate close indices
    deduped = [split_indices[0]]
    for idx in split_indices[1:]:
        if idx - deduped[-1] > 5:
            deduped.append(idx)
    split_indices = deduped
    split_indices.append(len(lines))

    all_items = []
    for i in range(len(split_indices) - 1):
        chunk_lines = lines[split_indices[i]:split_indices[i+1]]
        chunk_text = "\n".join(chunk_lines)
        items, _ = _regex_parse_single(chunk_text, chunk_lines)
        if items:
            all_items.extend(items)

    if not all_items:
        return [], "No product rows found. Set GEMINI_API_KEY for AI parsing."
    return all_items, None


# Kept for compatibility (also used internally)
def _regex_parse(text):
    return _regex_parse_text(text)


def _regex_parse_single(text, lines):
    """Regex parse a single invoice's text."""
    # Date
    m = re.search(r'\b(\d{1,2}[-](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-]\d{2,4})\b', text, re.IGNORECASE)
    date_val = m.group(1) if m else None
    if not date_val:
        m = re.search(r'\b(\d{1,2}[/]\d{1,2}[/]\d{4})\b', text)
        date_val = m.group(1) if m else None

    # Invoice number
    m = re.search(r'\b([A-Z]{1,6}[/\-]\d{2,4}[/\-]\d{2,4}[/\-]\d{1,6})\b', text)
    if not m:
        m = re.search(r'\b([A-Z]{1,6}[/\-]\d{1,6})\b', text)
    invoice_no = m.group(1) if m else 'UNKNOWN'

    # Customer: find "Buyer (Bill to)" section and pick up the company name
    customer = None
    SKIP = ['gstin', 'state', 'contact', 'udyam', 'iso', 'certified', 'invoice',
            'dated', 'voucher', 'mode', 'payment', 'buyer', 'bill to', 'dispatch',
            'destination', 'vehicle', 'lading', 'delivery', 'term', 'place', 'supply', 'road']
    buyer_section = False
    for line in lines:
        l, ll = line.strip(), line.strip().lower()
        if re.search(r'buyer\s*[\(\[]?\s*bill\s+to', ll) or 'bill to' in ll:
            buyer_section = True
            continue
        if buyer_section:
            if not l or len(l) < 5:
                continue
            if _is_likely_date(l) or _is_likely_number(l):
                continue
            if any(kw in ll for kw in SKIP):
                continue
            if re.match(r'^[A-Z]{2}\d{2}[A-Z]{2}\d{4}$', l):
                continue
            l = re.sub(r'\s+\d{1,2}[-/]\w{3,9}[-/]\d{2,4}\s*$', '', l).strip()
            if l and len(l) > 5:
                customer = l
                break

    if not customer:
        for line in lines:
            if re.search(r'\b(PVT|LTD|PRIVATE|LIMITED|INDUSTRIES|ENTERPRISES|ENGINEERING)\b',
                         line, re.IGNORECASE) and len(line.strip()) > 8:
                customer = line.strip()
                break

    # Items
    HEADER_KW = ['description', 'goods', 'hsn', 'sac', 'qty', 'quantity', 'amount', 'rate']
    hdr_idx = None
    for i, line in enumerate(lines):
        if sum(1 for kw in HEADER_KW if kw in line.lower()) >= 3:
            hdr_idx = i
            break

    items = []
    if hdr_idx is not None:
        for line in lines[hdr_idx + 1:]:
            l = line.strip()
            if not l:
                continue
            if any(kw in l.lower() for kw in SKIP_PRODUCT_KEYWORDS):
                continue
            if not re.search(r'\d{1,3}(?:,\d{3})*\.\d{2}\s*$', l):
                continue
            if re.match(r'^[\d\s%,\.]+$', l):
                continue

            # Parse right-to-left: Amount → Unit → Rate → Qty → Desc
            amount_m = re.search(r'([\d,]+\.\d{2})\s*$', l)
            if not amount_m:
                continue
            amount = float(amount_m.group(1).replace(',', ''))
            l = l[:amount_m.start()].strip()

            unit_m = re.search(r'\b(NOS|PCS|KG|MTR|SET|BOX|LTR|MTS|UNIT)\.?\s*$', l, re.IGNORECASE)
            unit = unit_m.group(1).upper() if unit_m else ''
            if unit_m:
                l = l[:unit_m.start()].strip()

            rate_m = re.search(r'([\d,]+\.\d{2})\s*$', l)
            if rate_m:
                l = l[:rate_m.start()].strip()

            qty_m = re.search(r'(\d+(?:\.\d+)?)\s*(NOS|PCS|KG|MTR|SET|BOX|LTR|MTS|UNIT)?\.?\s*$', l, re.IGNORECASE)
            qty = float(qty_m.group(1)) if qty_m else None
            if qty_m:
                l = l[:qty_m.start()].strip()
                if not unit and qty_m.group(2):
                    unit = qty_m.group(2).upper()

            l = re.sub(r'\s+\d{4,8}\s*$', '', l)
            desc = re.sub(r'^\d+\s+', '', l).strip()

            if desc and len(desc) > 3 and amount:
                items.append({'description': desc, 'quantity': qty, 'unit': unit, 'amount': amount})

    if not items:
        return [], "No product rows found."

    return [{'date': date_val, 'invoice_no': invoice_no, 'customer': customer or 'UNKNOWN',
             'product': it['description'], 'quantity': it['quantity'],
             'unit': it['unit'], 'amount': it['amount']} for it in items], None


# Alias for backward compat
_parse_single_invoice_regex = _regex_parse_single


# ─────────────────────────────────────────────────────────────
# MULTI-INVOICE INGESTION (entry point for app.py)
# ─────────────────────────────────────────────────────────────

def ingest_multiple_invoices(file_list, api_key=None):
    """
    Parse a list of (filename, bytes) tuples as GST invoices.
    Supports .pdf and visual .xlsx files (converted from PDFs).
    Returns (combined_df, list_of_errors).
    """
    all_rows, errors = [], []

    for fname, fbytes in file_list:
        fname_lower = fname.lower()
        if fname_lower.endswith(('.xlsx', '.xls')):
            # Use cell-position-aware openpyxl parser for visual Excel files
            rows, err = parse_excel_invoice(fbytes)
            if err:
                errors.append(f"{fname}: {err}")
            else:
                all_rows.extend(rows)
        else:
            # Handle PDFs: AI if key available, otherwise text regex
            rows, err = parse_gst_invoice_pdf(fbytes, api_key=api_key)
            if err:
                errors.append(f"{fname}: {err}")
            else:
                all_rows.extend(rows)

    if not all_rows:
        return pd.DataFrame(), errors

    df = pd.DataFrame(all_rows)
    df['date']     = pd.to_datetime(df['date'], errors='coerce', dayfirst=True)
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce')
    df['amount']   = pd.to_numeric(df['amount'],   errors='coerce')
    df = df.dropna(subset=['customer']).reset_index(drop=True)
    return df, errors
